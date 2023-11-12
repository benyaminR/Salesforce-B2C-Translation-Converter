import openpyxl
import requests
import json

# Constants
DEEPL_API_URL = "https://api-free.deepl.com/v2/translate"
DEEPL_API_KEY = "yourkeyhere"

# Load the Excel file
wb = openpyxl.load_workbook("translations_multi_lang.xlsx")


def translate_with_deepl(text, target_lang):
    payload = {"text": [text], "target_lang": target_lang}  # DeepL expects a list
    headers = {
        "Authorization": f"DeepL-Auth-Key {DEEPL_API_KEY}",
        "User-Agent": "YourApp/1.2.3",
        "Content-Type": "application/json",
    }

    response = requests.post(DEEPL_API_URL, headers=headers, data=json.dumps(payload))

    try:
        translation = response.json()["translations"][0]["text"]
        print(f"{text} translated to {target_lang} {translation}")
    except Exception as e:
        print(e)

    try:
        # Attempt to parse the JSON and extract the translation
        return response.json()["translations"][0]["text"]
    except Exception as e:
        print(f"Error for text '{text}' to language '{target_lang}':")
        print("HTTP Status Code:", response.status_code)
        print("Response Content:", response.content)
        print("Error:", str(e))
        return "Translation Error"  # Placeholder error text


# Loop through each sheet in the workbook
for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    headers = [cell.value for cell in ws[1]]

    # Go through each row in the worksheet
    for row in ws.iter_rows(min_row=2):
        id_cell = row[0]

        for idx, cell in enumerate(row[1:], start=1):
            if cell.value == "N/A":
                lang_code = headers[idx]
                source_translation = row[headers.index("de")].value
                if source_translation:  # Ensure there's content to translate
                    cell.value = translate_with_deepl(source_translation, lang_code)

# Save the updated Excel file
wb.save("translated_with_deepl.xlsx")

print("Translations completed and saved to translated_with_deepl.xlsx")
